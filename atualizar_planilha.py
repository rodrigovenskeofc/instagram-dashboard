"""
atualizar_planilha.py — Coleta os Reels do Instagram (@recorrenciadehonorarios),
preenche TODAS as colunas da aba "2026 Reels Org" e grava de volta no MESMO
arquivo .xlsx do Google Drive (preserva local e link, via Drive API files.update).

Modos:
  # Backfill / teste: preenche um intervalo a partir de uma linha inicial
  python atualizar_planilha.py --range 2026-06-16 2026-06-23 --start-row 60

  # Diario: preenche a proxima linha vazia com o Reel mais recente ainda nao registrado
  python atualizar_planilha.py --daily

  # Simula sem gravar no Drive (gera so a copia local)
  python atualizar_planilha.py --range 2026-06-16 2026-06-23 --start-row 60 --dry-run

Requisitos: .env (token IGAA) e gdrive_sa.json (chave da conta de servico) na mesma pasta.
Nada e enviado para fora alem da Graph API do Instagram e da Drive API do Google.
"""
import os, sys, json, io, argparse, urllib.request, urllib.parse, urllib.error
from datetime import datetime, timedelta, timezone, date
from pathlib import Path

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

PASTA   = Path(__file__).parent
FILE_ID = "1OspNVc8Sd-JvE1NHXM-l3-NWayG7dKfx"   # planilha no Drive
ABA     = "2026 Reels Org"
SA_JSON = PASTA / "gdrive_sa.json"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Proporcao para estimar Alcance Seguidores / Nao-Seguidores (P/Q),
# ja que a Graph API nao entrega mais o breakdown follow_type para reach.
# 0.255 = media historica observada na propria planilha (seguidores / alcance total).
RATIO_SEG = 0.255

# Regra do run diario: registrar o Reel postado "hoje - DELAY_DIAS" (maturacao da
# metrica). Rodrigo confirmou 3 dias (roda em 26/06 -> reel de 23/06).
DELAY_DIAS  = 3

BRT = timezone(timedelta(hours=-3))
DIAS_PT = ["segunda-feira","terça-feira","quarta-feira","quinta-feira",
           "sexta-feira","sábado","domingo"]

# Colunas (1-indexado) da aba "2026 Reels Org"
# P..S = split de Views por publico (lido da tela de insights, via navegador logado)
COL = dict(nr=1, distrib=2, seguidores=3, dif=4, atualizacao=5, data=6, dia=7,
           link=8, head=9, views=10, likes=11, coments=12, shares=13, saved=14,
           alcance=15)
# P..S (16..19) foram usadas pelo split/Views-IG via navegador, hoje REMOVIDO
# (a leitura logada travou a conta). API-only: essas colunas ficam em branco.

# ---------------- credenciais (.env local OU variaveis de ambiente na nuvem) ----------------
ENV_FILE = PASTA / ".env"

def load_env():
    env = {}
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            if "=" in line and not line.strip().startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

ENV   = load_env()
# token: prioriza .env local; na nuvem (GitHub Actions) vem de variavel de ambiente
TOKEN = ENV.get("INSTAGRAM_ACCESS_TOKEN") or os.environ.get("INSTAGRAM_ACCESS_TOKEN")
if not TOKEN:
    sys.exit("ERRO: INSTAGRAM_ACCESS_TOKEN ausente (.env ou variavel de ambiente).")
BASE  = "https://graph.instagram.com"

# ---------------- Instagram Graph ----------------
def ig_get(endpoint, params):
    params = dict(params); params["access_token"] = TOKEN
    url = f"{BASE}/{endpoint}?" + urllib.parse.urlencode(params)
    with urllib.request.urlopen(url, timeout=30) as r:
        return json.loads(r.read())

def refresh_token():
    """Renova o token IGAA (long-lived) e regrava no .env. Mantem a automacao
    perpetua: cada token vale ~60 dias e o refresh estende +60 a cada execucao.
    Best-effort: se falhar (ex.: token com <24h), apenas avisa e segue."""
    global TOKEN
    try:
        url = (f"{BASE}/refresh_access_token?grant_type=ig_refresh_token"
               f"&access_token={TOKEN}")
        with urllib.request.urlopen(url, timeout=30) as r:
            data = json.loads(r.read())
        novo = data.get("access_token")
        if novo and novo != TOKEN:
            p = PASTA / ".env"
            txt = p.read_text(encoding="utf-8").splitlines()
            out = []
            for line in txt:
                if line.strip().startswith("INSTAGRAM_ACCESS_TOKEN="):
                    out.append(f"INSTAGRAM_ACCESS_TOKEN={novo}")
                else:
                    out.append(line)
            p.write_text("\n".join(out) + "\n", encoding="utf-8")
            TOKEN = novo
            print(f"Token renovado (expira em ~{data.get('expires_in', 0)//86400} dias).")
        else:
            print("Token ainda valido (sem necessidade de troca).")
    except Exception as e:
        print(f"AVISO: nao consegui renovar o token ({e}). Seguindo com o atual.")

def get_profile():
    return ig_get("me", {"fields": "id,username,followers_count,media_count"})

def media_insights(media_id):
    out = {"reach": None, "saved": None, "shares": None,
           "views": None, "likes": None, "comments": None}
    try:
        d = ig_get(f"{media_id}/insights", {"metric": "reach"})
        node = d["data"][0]
        out["reach"] = (node.get("values") or [{}])[0].get("value")
        if out["reach"] is None:
            out["reach"] = node.get("total_value", {}).get("value")
    except Exception:
        pass
    for m in ["saved", "shares", "views", "likes", "comments"]:
        try:
            node = ig_get(f"{media_id}/insights", {"metric": m})["data"][0]
            val = (node.get("values") or [{}])[0].get("value")
            if val is None:
                val = node.get("total_value", {}).get("value")
            out[m] = val
        except Exception:
            pass
    return out

def fetch_reels(d1, d2):
    """Retorna lista de Reels (dicts) postados entre d1 e d2 (datas BRT), ordenados por data."""
    cutoff = datetime.combine(d1, datetime.min.time(), tzinfo=BRT) - timedelta(days=2)
    fields = ("id,caption,media_type,media_product_type,timestamp,"
              "permalink,like_count,comments_count")
    params = {"fields": fields, "limit": 50}
    items, ep = [], "me/media"
    while True:
        data = ig_get(ep, params)
        page = data.get("data", [])
        stop = False
        for it in page:
            items.append(it)
            ts = datetime.fromisoformat(it["timestamp"].replace("+0000", "+00:00"))
            if ts.astimezone(BRT) < cutoff:
                stop = True
        after = data.get("paging", {}).get("cursors", {}).get("after")
        if stop or not after:
            break
        params = {"after": after, "limit": 50, "fields": fields}

    reels = []
    for it in items:
        ts_brt = datetime.fromisoformat(it["timestamp"].replace("+0000", "+00:00")).astimezone(BRT)
        is_reel = it.get("media_product_type") == "REELS" or it.get("media_type") == "VIDEO"
        if is_reel and d1 <= ts_brt.date() <= d2:
            ins = media_insights(it["id"])
            head = (it.get("caption") or "").strip().split("\n")[0]
            reels.append({
                "id": it["id"], "post": ts_brt.date(),
                "link": it.get("permalink"), "head": head,
                "views": ins["views"],
                "likes": ins["likes"] if ins["likes"] is not None else it.get("like_count"),
                "coments": ins["comments"] if ins["comments"] is not None else it.get("comments_count"),
                "shares": ins["shares"], "saved": ins["saved"], "alcance": ins["reach"],
            })
    reels.sort(key=lambda x: x["post"])
    return reels

def followers_por_dia(d1, d2, total_now):
    """Reconstroi o total de seguidores no fim de cada dia e o ganho diario,
    a partir do insight follower_count (ganhos/dia) + total atual."""
    # A janela TEM que ir ate hoje: total_now e o valor de HOJE, entao para achar
    # o total numa data passada precisamos descontar TODOS os ganhos ate hoje.
    hoje_utc = datetime.now(timezone.utc).date()
    since = int(datetime.combine(d1 - timedelta(days=2), datetime.min.time(), tzinfo=timezone.utc).timestamp())
    until = int(datetime.combine(hoje_utc + timedelta(days=2), datetime.min.time(), tzinfo=timezone.utc).timestamp())
    gains = {}
    try:
        node = ig_get("me/insights",
                      {"metric": "follower_count", "period": "day",
                       "since": since, "until": until})["data"][0]
        for v in node["values"]:
            dt = datetime.fromisoformat(v["end_time"].replace("+0000", "+00:00")).date()
            gains[dt] = v.get("value") or 0
    except Exception:
        pass
    def total_no_fim(dia):
        return total_now - sum(g for dd, g in gains.items() if dd > dia)
    return {"total": total_no_fim, "dif": lambda dia: gains.get(dia, 0), "gains": gains}

# ---------------- Google Drive ----------------
def drive():
    scopes = ["https://www.googleapis.com/auth/drive"]
    if SA_JSON.exists():
        creds = service_account.Credentials.from_service_account_file(str(SA_JSON), scopes=scopes)
    elif os.environ.get("GDRIVE_SA"):
        info = json.loads(os.environ["GDRIVE_SA"])
        creds = service_account.Credentials.from_service_account_info(info, scopes=scopes)
    else:
        sys.exit("ERRO: credencial Google ausente (gdrive_sa.json ou variavel GDRIVE_SA).")
    return build("drive", "v3", credentials=creds, cache_discovery=False)

def drive_download(svc, dest):
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, svc.files().get_media(fileId=FILE_ID, supportsAllDrives=True))
    done = False
    while not done:
        _, done = dl.next_chunk()
    Path(dest).write_bytes(buf.getvalue())
    return len(buf.getvalue())

def drive_update(svc, src):
    media = MediaFileUpload(str(src), mimetype=XLSX_MIME, resumable=True)
    svc.files().update(fileId=FILE_ID, media_body=media, supportsAllDrives=True).execute()

# ---------------- Preenchimento (API-only) ----------------
def preencher_linha(ws, row, reel, segs, hoje):
    seg = segs["total"](reel["post"]); dif = segs["dif"](reel["post"])
    reach = reel["alcance"] or 0
    # A (nr) so se estiver vazia: continua a sequencia da linha de cima
    if ws.cell(row=row, column=COL["nr"]).value in (None, ""):
        prev = ws.cell(row=row - 1, column=COL["nr"]).value
        ws.cell(row=row, column=COL["nr"]).value = (int(prev) + 1) if isinstance(prev, (int, float)) else 1
    # B (distribuicao) e manual -> nao mexe
    ws.cell(row=row, column=COL["seguidores"]).value = seg
    ws.cell(row=row, column=COL["dif"]).value = dif
    c = ws.cell(row=row, column=COL["atualizacao"]); c.value = hoje; c.number_format = "DD/MM/YYYY"
    c = ws.cell(row=row, column=COL["data"]);        c.value = reel["post"]; c.number_format = "DD/MM/YYYY"
    ws.cell(row=row, column=COL["dia"]).value     = DIAS_PT[reel["post"].weekday()]
    ws.cell(row=row, column=COL["link"]).value    = reel["link"]
    ws.cell(row=row, column=COL["head"]).value    = reel["head"]
    # Views = metrica `views` da API = Visualizacoes do INSTAGRAM (sem Facebook).
    ws.cell(row=row, column=COL["views"]).value   = reel["views"]
    ws.cell(row=row, column=COL["likes"]).value   = reel["likes"]
    ws.cell(row=row, column=COL["coments"]).value = reel["coments"]
    ws.cell(row=row, column=COL["shares"]).value  = reel["shares"]
    ws.cell(row=row, column=COL["saved"]).value   = reel["saved"]
    ws.cell(row=row, column=COL["alcance"]).value = reach
    # API-only: limpa as colunas P..S (split/Views-IG do navegador, descontinuado).
    for c in (16, 17, 18, 19):
        ws.cell(row=row, column=c).value = None

def proxima_linha_vazia(ws, start=60):
    r = start
    while ws.cell(row=r, column=COL["data"]).value not in (None, ""):
        r += 1
    return r

# ---------------- Main ----------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--range", nargs=2, metavar=("INICIO", "FIM"))
    ap.add_argument("--start-row", type=int, default=60)
    ap.add_argument("--daily", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    # Renova o token so no run LOCAL (grava no .env). Na nuvem o token vem de
    # secret, renovado pelo job do dashboard; aqui apenas usamos.
    if not args.dry_run and ENV_FILE.exists():
        refresh_token()
    prof = get_profile()
    total_now = int(prof.get("followers_count") or 0)
    print(f"Conta: @{prof.get('username')} | Seguidores: {total_now}")
    hoje = datetime.now(BRT).date()

    if args.daily:
        # alvo: Reels postados ate (hoje - DELAY_DIAS).
        alvo_max = hoje - timedelta(days=DELAY_DIAS)
        print(f"Alvo: Reels postados ate {alvo_max:%d/%m/%Y} (hoje - {DELAY_DIAS} dias).")

        svc = None if args.dry_run else drive()
        src = PASTA / "planilha_drive_atual.xlsx"
        if svc:
            drive_download(svc, src)
        else:
            src = PASTA / "planilha_original.xlsx"
        wb = openpyxl.load_workbook(src); ws = wb[ABA]

        # ultima data de Reel ja registrada na secao 2026 (linha 60 em diante).
        # So registramos o que vier DEPOIS dela (nunca volta ao passado pre-teste),
        # ate hoje-DELAY. Se a automacao pular dias, recupera a partir daqui.
        datas = []
        for r in range(60, ws.max_row + 1):
            v = ws.cell(row=r, column=COL["data"]).value
            if isinstance(v, datetime): datas.append(v.date())
            elif isinstance(v, date):   datas.append(v)
        ultima = max(datas) if datas else alvo_max

        def norm(u): return str(u).split("?")[0].rstrip("/")
        ja = {norm(ws.cell(row=r, column=COL["link"]).value)
              for r in range(60, ws.max_row + 1)
              if ws.cell(row=r, column=COL["link"]).value}

        reels = fetch_reels(ultima, alvo_max)
        novos = [x for x in reels
                 if ultima <= x["post"] <= alvo_max and norm(x["link"]) not in ja]
        if not novos:
            print(f"Nenhum Reel novo (ultimo registrado: {ultima:%d/%m/%Y})."); return

        segs = followers_por_dia(min(x["post"] for x in novos), alvo_max, total_now)
        row = proxima_linha_vazia(ws, 60)
        for reel in novos:
            preencher_linha(ws, row, reel, segs, hoje)
            print(f"L{row} <- {reel['post']:%d/%m/%Y} | views={reel['views']} reach={reel['alcance']}")
            row += 1
        wb.save(src)
        if svc:
            drive_update(svc, src); print("GRAVADO no Drive (mesmo arquivo/link).")
        else:
            print(f"DRY-RUN: salvo so local em {src}")
        return

    if not args.range:
        ap.error("informe --range INICIO FIM (ou use --daily)")
    d1 = datetime.strptime(args.range[0], "%Y-%m-%d").date()
    d2 = datetime.strptime(args.range[1], "%Y-%m-%d").date()
    reels = fetch_reels(d1, d2)
    print(f"Reels encontrados {d1}..{d2}: {len(reels)}")
    segs = followers_por_dia(d1, d2, total_now)

    svc = None if args.dry_run else drive()
    src = PASTA / "planilha_drive_atual.xlsx"
    if svc:
        n = drive_download(svc, src); print(f"Baixou versao atual do Drive ({n} bytes).")
    else:
        src = PASTA / "planilha_original.xlsx"
        print("DRY-RUN: usando copia local planilha_original.xlsx")
    wb = openpyxl.load_workbook(src); ws = wb[ABA]

    row = args.start_row
    for reel in reels:
        preencher_linha(ws, row, reel, segs, hoje)
        seg = segs["total"](reel["post"])
        print(f"L{row} <- {reel['post']:%d/%m/%Y} {DIAS_PT[reel['post'].weekday()]:<13} "
              f"seg={seg} views={reel['views']} reach={reel['alcance']} "
              f"lk={reel['likes']} cm={reel['coments']} sh={reel['shares']} sv={reel['saved']}")
        row += 1

    out = PASTA / "planilha_TESTE_preenchida.xlsx"; wb.save(out)
    if svc:
        drive_update(svc, out)
        print(f"\nOK -> GRAVADO no Drive (mesmo arquivo/link): linhas {args.start_row}..{row-1}")
    else:
        print(f"\nDRY-RUN -> salvo so local em {out}")

if __name__ == "__main__":
    main()
